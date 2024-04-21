import io
import warnings
import cv2
import numpy as np
import pandas as pd
from PIL import Image
from openpyxl import load_workbook
from until import skylake

path = "/Users/mabowen/Downloads/皇宋、天圣、绍圣.xlsx"
warnings.simplefilter("ignore")
wb = load_workbook(path, data_only=True)
sheets = wb.sheetnames
client = skylake.Skylake('test')
wk_image = {}
for sheet in sheets:
    wk_sheet = wb[sheet]
    df = pd.read_excel(path, sheet_name=sheet)
    for image in wk_sheet._images:
        _, img_byte_array = cv2.imencode('.jpg',
                                         np.asarray(Image.open(io.BytesIO(image._data())).convert("RGB"))[:, :, ::-1])
        resp = client.image_binary_upload(img_byte_array.tobytes(), business_name="yzlzs", scene_name="imagecb")
        row = image.anchor._from.row - 1 if image.anchor._from.row > 1 else 0
        col = image.anchor._from.col
        key = f"{sheet}.{row}.{col}"
        if resp['code'] == 0:
            url = f"https://cdn.weipaitang.com{resp['data']['filename']}"
            df.iloc[row, col] = url
            print(f"{key}成功:{url}")
        else:
            print(f"{key}失败")
    df.to_excel(f"./{sheet}.xlsx", index=False)
